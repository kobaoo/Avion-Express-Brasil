# analytics.py — build 6 charts (matplotlib), interactive plotly, and Excel export
import os
import math
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import text
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule
import numpy as np

from config import engine

CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"
os.makedirs(CHARTS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

# ----------------------------
# Helpers
# ----------------------------
def run_query(sql: str) -> pd.DataFrame:
    with engine.connect() as conn:
        df = pd.read_sql(text(sql), conn)
    return df

def console_report(df: pd.DataFrame, chart_type: str, title: str):
    print(f"[{chart_type}] rows={len(df)} | {title}")

def save_png_current(fig_name: str):
    # Each chart its own figure, no seaborn, no explicit colors.
    plt.tight_layout()
    out = os.path.join(CHARTS_DIR, fig_name)
    plt.savefig(out, dpi=160, bbox_inches="tight")
    plt.close()
    print(f"Saved chart: {out}")

# ----------------------------
# Queries
# ----------------------------

# purchase by months
Q_LINE = """
WITH order_totals AS (
  SELECT o.order_id,
         DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
         SUM(oi.price + oi.freight_value) AS order_total
  FROM orders o
  JOIN order_items oi ON o.order_id = oi.order_id
  GROUP BY o.order_id, DATE_TRUNC('month', o.order_purchase_timestamp)
)
SELECT month, SUM(order_total) AS monthly_revenue
FROM order_totals
GROUP BY month
ORDER BY month;
"""

# payment types in orders
Q_PIE = """
SELECT op.payment_type,
       COUNT(*) AS cnt
FROM order_payments op
JOIN orders o ON op.order_id = o.order_id
JOIN order_items oi ON o.order_id = oi.order_id
GROUP BY op.payment_type
ORDER BY cnt DESC;
"""


Q_BAR = """
SELECT
    COALESCE(t.product_category_name_english, p.product_category_name) AS category,
    EXTRACT(YEAR FROM o.order_purchase_timestamp)::int AS year,
    ROUND(SUM(oi.price), 2) AS revenue
FROM order_items oi
JOIN orders o ON oi.order_id = o.order_id
JOIN products p ON p.product_id = oi.product_id
LEFT JOIN product_category_name_translation t
  ON t.product_category_name = p.product_category_name
WHERE EXTRACT(YEAR FROM o.order_purchase_timestamp) IN (2016,2017,2018)
GROUP BY category, year
ORDER BY SUM(oi.price) DESC
LIMIT 30;   -- 10 категорий * 3 года
    """

Q_BARH_REVIEWS = """
WITH scored AS (
  SELECT COALESCE(t.product_category_name_english, p.product_category_name) AS category,
         orv.review_score
  FROM order_reviews orv
  JOIN orders o ON orv.order_id = o.order_id
  JOIN order_items oi ON o.order_id = oi.order_id
  JOIN products p ON oi.product_id = p.product_id
  LEFT JOIN product_category_name_translation t
    ON t.product_category_name = p.product_category_name
  WHERE orv.review_score IS NOT NULL
)
SELECT category,
       ROUND(AVG(review_score)::NUMERIC, 3) AS avg_score,
       COUNT(*) AS n_reviews
FROM scored
GROUP BY category
HAVING COUNT(*) >= 50
ORDER BY avg_score DESC, n_reviews DESC
LIMIT 10;
"""
Q_SCATTER = """
WITH per_order AS (
  SELECT o.order_id,
         EXTRACT(EPOCH FROM (o.order_delivered_customer_date - o.order_purchase_timestamp))/86400.0 AS delivery_days,
         SUM(oi.price + oi.freight_value) AS order_total
  FROM orders o
  JOIN order_items oi ON o.order_id = oi.order_id
  JOIN order_payments op ON op.order_id = o.order_id
  WHERE o.order_delivered_customer_date IS NOT NULL
  GROUP BY o.order_id, delivery_days
)
SELECT * FROM per_order
WHERE delivery_days BETWEEN 0 AND 60;
"""

Q_ORDERS_BY_YEARS = """
WITH monthly AS (
    SELECT
        DATE_TRUNC('month', order_purchase_timestamp) AS month_start,
        EXTRACT(YEAR FROM order_purchase_timestamp)::int AS year,
        COUNT(*)::int AS order_count
    FROM orders
    WHERE order_purchase_timestamp IS NOT NULL
        AND EXTRACT(YEAR FROM order_purchase_timestamp) BETWEEN 2016 AND 2018
    GROUP BY 1, 2
)
SELECT month_start::date AS month_start, year, order_count
FROM monthly
ORDER BY month_start, year;
"""

# ----------------------------
# Charts (matplotlib)
# ----------------------------
def chart_line_monthly_revenue():
    df = run_query(Q_LINE)
    console_report(df, "line", "Monthly revenue trend")
    plt.figure()
    plt.plot(df["month"], df["monthly_revenue"], marker="o")
    plt.title("Monthly Revenue Trend")
    plt.xlabel("Month")
    plt.ylabel("Revenue")
    plt.xticks(ticks=df["month"],rotation=90)
    save_png_current("01_line_monthly_revenue.png")
    return df

def chart_pie_payment_share():
    df = run_query(Q_PIE)
    console_report(df, "pie", "Payment method share (count of payments)")

    sizes = df["cnt"].values
    labels = df["payment_type"].values
    colors = plt.cm.Set3(np.linspace(0, 1, len(labels)))

    explode = [0.05] + [0]*(len(labels)-1)

    fig, ax = plt.subplots(figsize=(10, 8))
    wedges, texts, autotexts = ax.pie(
        sizes,
        labels=labels,
        autopct='%1.1f%%',
        colors=colors,
        explode=explode,
        shadow=True,
        startangle=90
    )

    for autotext in autotexts:
        autotext.set_color('black')
        autotext.set_weight('bold')

    ax.set_title('Payment Method Share', fontsize=16, pad=20)
    plt.tight_layout()

    save_png_current("02_pie_payment_share.png")
    return df

def chart_bar_top_categories_revenue():
    df = run_query(Q_BAR)
    console_report(df, "bar", "Top-10 Categories by Revenue, 2016–2018")

    # берём 10 самых прибыльных категорий суммарно
    top10 = (
        df.groupby("category")["revenue"]
        .sum()
        .sort_values(ascending=False)
        .head(10)
        .index
    )
    df = df[df["category"].isin(top10)]

    # строим группированную диаграмму
    plt.figure(figsize=(10, 6))
    years = [2016, 2017, 2018]
    width = 0.25
    x = np.arange(len(top10))
    colors = {2016: "skyblue", 2017: "orange", 2018: "green"}

    for i, y in enumerate(years):
        revs = [
            df[(df["category"] == cat) & (df["year"] == y)]["revenue"].sum()
            for cat in top10
        ]
        plt.bar(x + i * width - width, revs, width, label=str(y), color=colors[y])

    plt.title("Top-10 Categories by Revenue (2016–2018)")
    plt.xlabel("Category")
    plt.ylabel("Revenue")
    plt.xticks(x, top10, rotation=45, ha="right")
    plt.legend(title="Year")
    save_png_current("03_bar_top_categories_revenue_by_year.png")
    return df


def chart_barh_avg_review_by_category():
    df = run_query(Q_BARH_REVIEWS)
    console_report(df, "barh", "Top-10 Categories by Avg Review Score")
    plt.figure()
    colors = plt.cm.Set3(np.linspace(0, 1, len(df["category"])))
    plt.barh(df["category"], df["avg_score"], color=colors)
    plt.title("Top-10 Categories by Avg Review Score (n>=50)")
    plt.xlabel("Average Score")
    plt.ylabel("Category")
    plt.gca().invert_yaxis()
    save_png_current("04_barh_avg_review_score.png")
    return df

def chart_hist_order_total_distribution():
    df = run_query(Q_ORDERS_BY_YEARS)
    df['month_num'] = pd.to_datetime(df['month_start']).dt.month
    pivot = df.pivot_table(index='month_num', columns='year', values='order_count', aggfunc='sum').fillna(0).astype(int)
    # Ensure rows for all 12 months exist
    pivot = pivot.reindex(range(1,13), fill_value=0)    

    plt.figure(figsize=(12, 6))
    x = pivot.index.values 
    for i, year in enumerate(sorted(pivot.columns)):
        plt.bar(x, pivot[year].values, alpha=0.3, label=str(year))

    plt.title("Monthly Order Counts by Year (2016–2018)")
    plt.xlabel("Month")
    plt.ylabel("Orders")
    plt.xticks(ticks=range(1,13), labels=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"])
    plt.legend(title="Year")
    plt.tight_layout()
    save_png_current("05_hist_order_by_years.png")

def chart_scatter_delivery_vs_total():
    df = run_query(Q_SCATTER)
    console_report(df, "scatter", "Delivery time (days) vs Order total")
    plt.figure()
    plt.scatter(df["order_total"], df["delivery_days"], s=10, color='coral', alpha=0.6)
    plt.title("Delivery Time vs Order Total (0-60 days)")
    plt.xlabel("Order Total")
    plt.ylabel("Delivery Days")
    save_png_current("06_scatter_delivery_vs_total.png")
    return df
def interactive_time_slider():
    sql = """
    SET search_path TO olist, public;

    WITH orders_by_state AS (
      SELECT
        DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
        c.customer_state AS state
      FROM orders o
      JOIN customers c ON c.customer_id = o.customer_id
      WHERE o.order_purchase_timestamp IS NOT NULL
        AND c.customer_state IS NOT NULL
    ),
    top5 AS (
      SELECT state, COUNT(*) AS cnt
      FROM orders_by_state
      GROUP BY state
      ORDER BY cnt DESC
      LIMIT 5
    )
    SELECT
      obs.month,
      obs.state,
      COUNT(*)::int AS orders_count
    FROM orders_by_state obs
    JOIN top5 t ON t.state = obs.state
    GROUP BY obs.month, obs.state
    ORDER BY obs.month, obs.state;
    """

    df = run_query(sql)
    if df.empty:
        print("No data for plot.")
        return {"data": df, "fig": None}

    # ----- аккуратно выровняем месяцы, чтобы не было разрывов -----
    df["month"] = pd.to_datetime(df["month"]).dt.to_period("M").dt.to_timestamp()
    # полный месячный диапазон
    all_months = pd.period_range(df["month"].min(), df["month"].max(), freq="M").to_timestamp()
    states = df["state"].unique()

    # заполнить пропуски нулями для каждого штата
    filled = (
        df.set_index(["state", "month"])
          .reindex(pd.MultiIndex.from_product([states, all_months], names=["state","month"]))
          .fillna(0)
          .reset_index()
    )
    filled["orders_count"] = filled["orders_count"].astype(int)

    # Создаем накопительные данные для анимации
    filled_sorted = filled.sort_values(['state', 'month'])
    
    # Для каждого месяца показываем все данные до этого месяца включительно
    animation_frames = []
    for i, month in enumerate(sorted(filled['month'].unique())):
        frame_data = filled_sorted[filled_sorted['month'] <= month].copy()
        frame_data['frame'] = month
        animation_frames.append(frame_data)
    
    animated_data = pd.concat(animation_frames)

    # ----- сам график с анимацией по месяцам -----
    fig = px.line(
        animated_data,
        x="month",
        y="orders_count", 
        color="state",
        markers=True,
        template="plotly_white",
        title="Monthly Orders by Top 5 Customer States",
        animation_frame="frame",
        range_x=[filled['month'].min(), filled['month'].max()],
        range_y=[0, filled["orders_count"].max() * 1.1]
    )
    
    fig.update_traces(mode="lines+markers", line=dict(width=2))
    fig.update_layout(
        xaxis_title="Month",
        yaxis_title="Orders",
        legend_title="State", 
        hovermode="x unified"
    )

    fig.show()
    return {"data": filled, "fig": fig}


# ----------------------------
# Excel export with formatting
# ----------------------------
def export_to_excel(dfs: dict, filename: str):
    """
    dfs: {"sheet_name": DataFrame, ...}
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    import numbers
    import datetime as dt

    path = os.path.join(EXPORTS_DIR, filename)

    def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        # убрать tz у всех datetime колонок
        for c in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[c]):
                # если tz-aware -> сделать naive
                try:
                    df[c] = pd.to_datetime(df[c], utc=True).dt.tz_convert(None)
                except Exception:
                    df[c] = pd.to_datetime(df[c]).dt.tz_localize(None)
            # привести category к строке
            if pd.api.types.is_categorical_dtype(df[c]):
                df[c] = df[c].astype(str)
        return df

    # 1) записываем чистые данные
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            clean = _sanitize_df(df)
            clean.to_excel(writer, sheet_name=sheet, index=False)

    # 2) пост-стилизация
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Freeze только шапку
        ws.freeze_panes = "A2"

        # AutoFilter по использованному диапазону
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        # Жирная шапка
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Авто-ширина колонок (аккуратно, с ограничением)
        max_width = 40
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws[f"{col_letter}{row}"].value
                # приблизительная длина
                if v is None:
                    l = 0
                elif isinstance(v, (dt.date, dt.datetime)):
                    l = len(v.isoformat())
                else:
                    l = len(str(v))
                if l > max_len:
                    max_len = l
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

        # Градиент ТОЛЬКО по числовым колонкам (и где есть >1 значения)
        for col_idx in range(1, ws.max_column + 1):
            # пробуем найти хотя бы одно число в колонке
            values = []
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=col_idx).value
                if isinstance(v, numbers.Number):
                    values.append(v)
                # игнорируем даты/текст/None
            if len(values) >= 2:  # достаточно точек для шкалы
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",  # красный
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",  # жёлтый
                    end_type="max", end_color="FF00AA00"  # зелёный
                )
                ws.conditional_formatting.add(rng, rule)

    wb.save(path)
    total_rows = sum(len(df) for df in dfs.values())
    print(f'Created file {os.path.basename(path)}, {len(dfs)} sheets, {total_rows} rows at {path}')
    return path

# ----------------------------
# Main
# ----------------------------
def main():
    # 6 charts
    df_line = chart_line_monthly_revenue()
    df_pie = chart_pie_payment_share()
    df_bar = chart_bar_top_categories_revenue()
    df_barh = chart_barh_avg_review_by_category()
    hist = chart_hist_order_total_distribution()
    df_scatter = chart_scatter_delivery_vs_total()

    # interactive
    interactive_time_slider() 

    export_to_excel(
        {
            "monthly_revenue": df_line,
            "payment_share": df_pie,
            "top_categories_revenue": df_bar,
            "avg_review_by_category": df_barh,
            "delivery_vs_total": df_scatter,
        },
        f"olist_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

if __name__ == "__main__":
    main()
